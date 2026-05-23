import io
import json
import os
import uuid
from datetime import datetime
import logging
import threading
import pandas as pd
import smtplib
import functools
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
from flask import Flask, flash, redirect, jsonify, render_template, request, send_file, session, url_for
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename

from utils.validation import validate_estudio
from agents.analyzer import analizar_imagenes
from utils.backend import generar_reporte_pdf, guardar_archivo_local, guardar_analisis_json
from utils.supabase_client import (
    get_usuario_por_cedula,
    crear_usuario,
    actualizar_usuario,
    get_paciente_por_cedula,
    crear_paciente,
    actualizar_paciente,
    crear_o_actualizar_estudio,
    crear_imagen_medica,
    crear_linea_analisis,
    crear_prediccion_ia,
    crear_reporte_medico,
    obtener_importaciones_excel,
    obtener_reportes_medicos,
    obtener_reporte_medico_por_id,
    eliminar_resultado_completo,
    obtener_prediccion_por_id,
    obtener_linea_por_id,
    obtener_imagen_por_id,
    obtener_estudio_por_cedula,
    obtener_lineas_por_cedula,
    obtener_reportes_completos_por_cedula,
    eliminar_importacion_excel,
    listar_todos_usuarios,
    actualizar_rol_usuario,
    eliminar_usuario_completo,
)

rq_queue = None
try:
    import importlib
    redis = importlib.import_module('redis')
    rq_mod = importlib.import_module('rq')
    Queue = getattr(rq_mod, 'Queue')
    redis_url = os.getenv('REDIS_URL', 'redis://localhost:6379')
    redis_conn = redis.from_url(redis_url)
    rq_queue = Queue('mammo', connection=redis_conn)
except Exception:
    rq_queue = None

app = Flask(__name__, template_folder='templates', static_folder='static')
app.secret_key = os.getenv('SECRET_KEY', 'mammo2026-secret')


def login_required(f):
    @functools.wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('usuario'):
            flash('Debes iniciar sesion para acceder a esta seccion.', 'warning')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated


def medico_required(f):
    @functools.wraps(f)
    def decorated(*args, **kwargs):
        u = session.get('usuario')
        if not u:
            flash('Debes iniciar sesion.', 'warning')
            return redirect(url_for('login'))
        if u.get('rol', '').upper() != 'MEDICO':
            flash('Acceso restringido a medicos.', 'danger')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated


@app.context_processor
def inject_usuario():
    return {'current_user': session.get('usuario')}
app.config['MAX_CONTENT_LENGTH'] = 80 * 1024 * 1024

ALLOWED_IMAGE = {'png', 'jpg', 'jpeg', 'dcm', 'tiff', 'tif', 'bmp', 'gif'}

UPLOADS_DIR = os.path.join(os.getcwd(), 'uploads')
IMAGES_DIR = os.path.join(UPLOADS_DIR, 'images')
REPORTS_DIR = os.path.join(UPLOADS_DIR, 'reports')
TMP_DIR = os.path.join(UPLOADS_DIR, 'tmp')
for folder in [UPLOADS_DIR, IMAGES_DIR, REPORTS_DIR, TMP_DIR]:
    os.makedirs(folder, exist_ok=True)


def guardar_imagen_temporal(bytes_archivo, nombre_archivo):
    extension = os.path.splitext(nombre_archivo)[1].lower().lstrip('.') or 'png'
    nombre = f"{uuid.uuid4().hex}.{extension}"
    ruta = os.path.join(TMP_DIR, nombre)
    os.makedirs(os.path.dirname(ruta), exist_ok=True)
    with open(ruta, 'wb') as archivo:
        archivo.write(bytes_archivo)
    return ruta


def limpiar_archivos_temporales(rutas):
    for ruta in rutas:
        try:
            if ruta and os.path.exists(ruta):
                os.remove(ruta)
        except Exception:
            pass


def determinar_tipo_imagen(nombre_archivo):
    extension = os.path.splitext(nombre_archivo)[1].lower()
    if extension == '.dcm':
        return 'DICOM'
    if extension in {'.png', '.jpg', '.jpeg', '.tiff', '.tif', '.bmp', '.gif'}:
        return 'TERMOGRAFIA'
    return 'OTRO'


def format_history_records(records):
    formatted = []
    for record in records:
        if record.get('file_name') or record.get('nombre_archivo'):
            meta = record.get('import_metadata') or record.get('metadatos_importacion') or {}
            training = meta.get('training_history', [])
            formatted.append({
                'Fecha': record.get('imported_at') or record.get('importado_en') or record.get('created_at', ''),
                'Run': record.get('file_name') or record.get('nombre_archivo', ''),
                'Modelo': meta.get('model_name', meta.get('model', 'InceptionResNetV2')),
                'Epochs': record.get('total_records_processed') or record.get('total_registros_entrenamiento', ''),
                'HistorialEntrenamiento': json.dumps(training, ensure_ascii=False),
                'StoragePath': record.get('storage_path') or record.get('ruta_almacenamiento_xlsx', ''),
                'CedulaUsuario': record.get('cedula_usuario') or record.get('usuario_cedula', ''),
            })
        else:
            study = record.get('study_id') or record.get('estudios') or {}
            prediction = record.get('prediction_id') or record.get('predicciones_ia') or {}
            patient_id = study.get('cedula_paciente', '') if isinstance(study, dict) else ''
            formatted.append({
                'Fecha': record.get('created_at') or record.get('creado_en', ''),
                'Paciente': patient_id,
                'Estudio': study.get('uid_instancia_estudio', '') if isinstance(study, dict) else '',
                'Tipo predicción': prediction.get('patologia_predicha') or prediction.get('predicted_pathology', '') if isinstance(prediction, dict) else '',
                'Probabilidad malignidad': prediction.get('probabilidad_malignidad') or prediction.get('malignancy_probability', '') if isinstance(prediction, dict) else '',
                'Notas clínicas': record.get('notas_clinicas') or record.get('clinical_notes', ''),
                'PDF': record.get('ruta_pdf_almacenamiento') or record.get('pdf_storage_path', ''),
            })
    return formatted


@app.route('/login', methods=['GET', 'POST'])
def login():
    if session.get('usuario'):
        return redirect(url_for('index'))
    if request.method == 'POST':
        cedula = request.form.get('cedula', '').strip()
        contrasena = request.form.get('contrasena', '').strip()
        if not cedula or not contrasena:
            flash('Ingrese cédula y contraseña.', 'warning')
            return redirect(url_for('login'))
        usuario = get_usuario_por_cedula(cedula)
        if not usuario or not isinstance(usuario, dict) or usuario.get('error'):
            flash('Cédula no encontrada. Verifique sus datos.', 'danger')
            return redirect(url_for('login'))
        if not check_password_hash(usuario.get('contrasena_hash', ''), contrasena):
            flash('Contraseña incorrecta.', 'danger')
            return redirect(url_for('login'))
        session['usuario'] = {
            'cedula': usuario.get('cedula'),
            'nombre_completo': usuario.get('nombre_completo'),
            'rol': usuario.get('rol', 'PACIENTE'),
        }
        flash('Bienvenido, ' + usuario.get('nombre_completo', '') + '.', 'success')
        return redirect(url_for('index'))
    return render_template('login.html', active='login')


@app.route('/logout')
def logout():
    session.clear()
    flash('Sesión cerrada correctamente.', 'info')
    return redirect(url_for('login'))


@app.route('/')
def index():
    return render_template('index.html', active='home')


@app.route('/verificar_cedula')
def verificar_cedula():
    cedula = request.args.get('cedula', '').strip()
    if not cedula:
        return jsonify({'existe': False})
    usuario = get_usuario_por_cedula(cedula)
    existe = bool(usuario and isinstance(usuario, dict) and not usuario.get('error'))
    return jsonify({'existe': existe})


@app.route('/obtener_perfil/<cedula>')
def obtener_perfil(cedula):
    usuario = get_usuario_por_cedula(cedula)
    if not usuario or not isinstance(usuario, dict) or usuario.get('error'):
        return jsonify({'error': 'No encontrado'}), 404
    paciente = get_paciente_por_cedula(cedula) or {}
    return jsonify({'usuario': usuario, 'paciente': paciente})


@app.route('/registro', methods=['GET', 'POST'])
def registro():
    usuario_sesion = session.get('usuario', {})
    rol_sesion = usuario_sesion.get('rol', '')

    if request.method == 'POST':
        nombre_completo = request.form.get('nombre_completo', '').strip()
        cedula          = request.form.get('cedula', '').strip()
        correo          = request.form.get('correo', '').strip()
        edad            = request.form.get('edad', '').strip()
        telefono        = request.form.get('telefono', '').strip()
        sexo            = request.form.get('sexo', '').strip()
        direccion       = request.form.get('direccion', '').strip()
        fecha_nacimiento = request.form.get('fecha_nacimiento', '').strip()
        contrasena      = request.form.get('contrasena', '').strip()

        rol = request.form.get('rol', 'PACIENTE').strip() if rol_sesion == 'MEDICO' else 'PACIENTE'
        if rol not in ('PACIENTE', 'MEDICO'):
            rol = 'PACIENTE'

        if not nombre_completo or not cedula or not correo:
            flash('Nombre, cédula y correo son obligatorios.', 'warning')
            return redirect(url_for('registro'))

        usuario_existente = get_usuario_por_cedula(cedula)
        es_actualizar = (
            usuario_existente and
            isinstance(usuario_existente, dict) and
            not usuario_existente.get('error')
        )

        if es_actualizar:
            cedula_sesion = usuario_sesion.get('cedula', '')
            if rol_sesion != 'MEDICO' and cedula != cedula_sesion:
                flash('No tiene permiso para modificar otros usuarios.', 'danger')
                return redirect(url_for('registro'))
            datos_update = {
                'nombre_completo': nombre_completo,
                'correo': correo,
                'edad': int(edad) if edad else None,
                'telefono': telefono,
                'direccion': direccion,
            }
            if rol_sesion == 'MEDICO':
                datos_update['rol'] = rol
            if contrasena:
                datos_update['contrasena_hash'] = generate_password_hash(contrasena)
            actualizar_usuario(cedula, datos_update)
            actualizar_paciente(cedula, {
                'sexo': sexo if sexo in ('O', 'F', 'M') else 'F',
                'fecha_nacimiento': fecha_nacimiento if fecha_nacimiento else None,
            })
            flash('Datos actualizados con éxito.', 'success')
            return redirect(url_for('registro'))
        else:
            if not contrasena:
                flash('La contraseña es obligatoria para nuevos registros.', 'warning')
                return redirect(url_for('registro'))
            contrasena_hash = generate_password_hash(contrasena)
            nuevo_usuario = {
                'nombre_completo': nombre_completo,
                'cedula': cedula,
                'correo': correo,
                'edad': int(edad) if edad else None,
                'telefono': telefono,
                'rol': rol,
                'esta_activo': True,
                'direccion': direccion,
                'contrasena_hash': contrasena_hash,
            }
            usuario_creado = crear_usuario(nuevo_usuario)
            if usuario_creado and not (isinstance(usuario_creado, dict) and usuario_creado.get('error')):
                crear_paciente({
                    'cedula': cedula,
                    'sexo': sexo if sexo in ('O', 'F', 'M') else 'F',
                    'fecha_nacimiento': fecha_nacimiento if fecha_nacimiento else None,
                })
                flash('Registro completado con éxito.', 'success')
                return redirect(url_for('index'))
            else:
                flash('Error al registrar el usuario. Intente de nuevo.', 'danger')
                return redirect(url_for('registro'))

    usuario_data  = {}
    paciente_data = {}
    if usuario_sesion:
        cedula_sesion = usuario_sesion.get('cedula', '')
        u = get_usuario_por_cedula(cedula_sesion)
        p = get_paciente_por_cedula(cedula_sesion) or {}
        if u and isinstance(u, dict) and not u.get('error'):
            usuario_data  = u
            paciente_data = p
    return render_template('registro.html', active='registro', usuario_data=usuario_data, paciente_data=paciente_data)


@app.route('/upload', methods=['GET', 'POST'])
@login_required
@medico_required
def upload():
    scan_preview = session.get('scan_preview')
    if request.method == 'POST':
        cedula = request.form.get('cedula', '').strip()
        notas_clinicas = request.form.get('notas_clinicas', '').strip()
        archivos = request.files.getlist('archivos[]')

        if not cedula:
            flash('La cédula es obligatoria.', 'warning')
            return redirect(url_for('upload'))

        usuario_existente = get_usuario_por_cedula(cedula)
        if not usuario_existente or (isinstance(usuario_existente, dict) and usuario_existente.get('error')):
            flash('La cédula ingresada no existe en el sistema. Registre primero al usuario.', 'danger')
            return redirect(url_for('upload'))

        paciente_existente = get_paciente_por_cedula(cedula) or {}

        if not archivos or len(archivos) == 0:
            flash('Debe cargar al menos una imagen.', 'warning')
            return redirect(url_for('upload'))

        bytes_dicom = []
        bytes_termograficos = []
        temp_images = []
        for archivo in archivos:
            if not archivo or archivo.filename == '':
                continue
            nombre_archivo = secure_filename(archivo.filename)
            contenido = archivo.read()
            tipo = determinar_tipo_imagen(nombre_archivo)
            if tipo == 'DICOM':
                bytes_dicom.append(contenido)
            else:
                bytes_termograficos.append(contenido)
            ruta_temporal = guardar_imagen_temporal(contenido, nombre_archivo)
            temp_images.append({
                'temp_path': ruta_temporal,
                'filename': nombre_archivo,
                'tipo_imagen': tipo,
                'uid_instancia_sop': str(uuid.uuid4()),
                'metadatos': {'nombre_original': nombre_archivo, 'tamano_bytes': len(contenido)},
            })

        if not temp_images:
            flash('No se detectaron archivos válidos para escanear.', 'warning')
            return redirect(url_for('upload'))

        resultado = analizar_imagenes(bytes_dicom, bytes_termograficos)
        resultado['patient'] = {
            'cedula': cedula,
            'correo': usuario_existente.get('correo', ''),
            'nombre_completo': usuario_existente.get('nombre_completo', f'Paciente {cedula}'),
            'fecha_nacimiento': paciente_existente.get('fecha_nacimiento'),
            'sexo': paciente_existente.get('sexo', 'F'),
        }
        resultado['study'] = {
            'uid_instancia_estudio': str(uuid.uuid4()),
            'fecha_estudio': datetime.utcnow().date().isoformat(),
            'descripcion': 'Escaneo de imágenes médicas con IA',
        }

        scan_preview = {
            'patient': resultado['patient'],
            'study': resultado['study'],
            'analisis': resultado['analisis'],
            'summary': resultado['summary'],
            'training_history': resultado.get('training_history', []),
            'temp_images': temp_images,
            'notas_clinicas': notas_clinicas,
            'created_at': datetime.utcnow().isoformat(),
        }
        session['scan_preview'] = scan_preview
        return render_template('upload.html', active='upload', scan_preview=scan_preview)

    last_result = None
    if session.get('last_analysis'):
        try:
            last_result = json.loads(session['last_analysis'])
        except Exception:
            last_result = None
    return render_template('upload.html', active='upload', last_result=last_result, scan_preview=scan_preview)


@app.route('/upload/guardar', methods=['POST'])
def guardar_scan():
    scan_preview = session.pop('scan_preview', None)
    if not scan_preview:
        flash('No hay escaneo pendiente para guardar.', 'warning')
        return redirect(url_for('upload'))

    patient = scan_preview['patient']
    cedula = patient['cedula']
    notas_clinicas = scan_preview.get('notas_clinicas', '')

    usuario = get_usuario_por_cedula(cedula)
    if not usuario or (isinstance(usuario, dict) and usuario.get('error')):
        flash('La cédula no existe en el sistema. No se puede guardar.', 'danger')
        limpiar_archivos_temporales([item['temp_path'] for item in scan_preview.get('temp_images', [])])
        return redirect(url_for('upload'))

    paciente = get_paciente_por_cedula(cedula)
    if paciente:
        actualizar_paciente(cedula, {
            'fecha_nacimiento': patient.get('fecha_nacimiento'),
            'sexo': patient.get('sexo', 'F'),
        })
    else:
        crear_paciente({
            'cedula': cedula,
            'fecha_nacimiento': patient.get('fecha_nacimiento'),
            'sexo': patient.get('sexo', 'F'),
        })

    estudio_payload = {
        'cedula_paciente': cedula,
        'uid_instancia_estudio': scan_preview['study']['uid_instancia_estudio'],
        'fecha_estudio': scan_preview['study']['fecha_estudio'],
        'descripcion': scan_preview['study']['descripcion'],
    }
    is_valid_est, est_msg = validate_estudio(estudio_payload)
    if not is_valid_est:
        flash(est_msg or 'Datos de estudio inválidos.', 'danger')
        limpiar_archivos_temporales([item['temp_path'] for item in scan_preview.get('temp_images', [])])
        return redirect(url_for('upload'))

    estudio = crear_o_actualizar_estudio(estudio_payload)
    if isinstance(estudio, dict) and estudio.get('error'):
        flash('No se pudo crear el estudio en la base de datos.', 'danger')
        limpiar_archivos_temporales([item['temp_path'] for item in scan_preview.get('temp_images', [])])
        return redirect(url_for('upload'))

    medical_records = []
    for item, temp in zip(scan_preview['analisis'], scan_preview['temp_images']):
        try:
            with open(temp['temp_path'], 'rb') as f:
                contenido = f.read()
        except Exception:
            contenido = b''
        extension = os.path.splitext(temp['temp_path'])[1].lstrip('.') or 'png'
        storage_path = guardar_archivo_local(contenido, 'images', extension)
        if os.path.exists(temp['temp_path']):
            os.remove(temp['temp_path'])

        imagen = crear_imagen_medica({
            'cedula_paciente': cedula,
            'tipo_imagen': temp['tipo_imagen'],
            'ruta_almacenamiento': storage_path,
            'uid_instancia_sop': temp['uid_instancia_sop'],
            'metadatos': temp.get('metadatos', {}),
        })
        if isinstance(imagen, dict) and imagen.get('error'):
            logging.warning('crear_imagen_medica error: %s', imagen.get('error'))
            flash('No se pudo guardar una imagen en la base de datos.', 'danger')
            limpiar_archivos_temporales([t['temp_path'] for t in scan_preview.get('temp_images', []) if os.path.exists(t.get('temp_path', ''))])
            return redirect(url_for('upload'))
        medical_records.append((imagen, item, storage_path))

    predictions = []
    for imagen, item, storage_path in medical_records:
        score_riesgo = item.get('score_riesgo', 0)
        agentes = item.get('detalles_agentes', [])
        boxes_cnn = []
        centroides_kmeans = []
        for ag in agentes:
            if ag.get('agente') == 'CNN':
                boxes_cnn = ag.get('boxes', [])
            elif ag.get('agente') == 'IsolationForest':
                centroides_kmeans = ag.get('boxes_anomalias', [])

        linea = crear_linea_analisis({
            'cedula_paciente': cedula,
            'imagen_id': imagen.get('id'),
            'estado': 'COMPLETADO',
            'parametros_filtros_reduccion_ruido': {'metodo': 'nl_means', 'patch_size': 5, 'patch_distance': 6},
            'registros_preprocesamiento': 'Imagen procesada y normalizada con preprocesamiento adaptativo',
            'tiempo_ejecucion_segundos': 2.5,
        })
        if isinstance(linea, dict) and linea.get('error'):
            logging.warning('crear_linea_analisis error: %s', linea.get('error'))
            flash('No se pudo crear la línea de análisis.', 'danger')
            return redirect(url_for('upload'))

        patologia = 'MALIGNA' if score_riesgo >= 50 else 'BENIGNA' if score_riesgo >= 20 else 'NORMAL'
        prediccion = crear_prediccion_ia({
            'cedula_paciente': cedula,
            'linea_analisis_id': linea.get('id'),
            'modelo_nombre': 'MammoScan Multi-Agent Ensemble',
            'probabilidad_malignidad': round(score_riesgo / 100, 4),
            'patologia_predicha': patologia,
            'cajas_delimitadoras_cnn': boxes_cnn,
            'centroides_kmeans': centroides_kmeans,
            'ruta_mapa_calor_resnet': storage_path,
        })
        if isinstance(prediccion, dict) and prediccion.get('error'):
            logging.warning('crear_prediccion_ia error: %s', prediccion.get('error'))
            flash('No se pudo guardar la predicción en la base de datos.', 'danger')
            return redirect(url_for('upload'))
        predictions.append(prediccion)

    patient_para_pdf = {
        **patient,
        'nombre_completo': usuario.get('nombre_completo', patient.get('nombre_completo', '')),
        'correo': usuario.get('correo', patient.get('correo', '')),
    }
    report_bytes = generar_reporte_pdf({
        'analisis': scan_preview['analisis'],
        'summary': scan_preview['summary'],
        'training_history': scan_preview.get('training_history', []),
        'study': scan_preview['study'],
        'notas_clinicas': notas_clinicas,
    }, patient_para_pdf)
    report_path = guardar_archivo_local(report_bytes, 'reports', 'pdf')

    reporte = crear_reporte_medico({
        'cedula_paciente': cedula,
        'prediccion_id': predictions[0].get('id') if predictions else None,
        'ruta_pdf_almacenamiento': report_path,
        'notas_clinicas': notas_clinicas if notas_clinicas else 'Informe generado automáticamente por MammoScan AI.',
        'esta_verificado': False,
    })
    if isinstance(reporte, dict) and reporte.get('error'):
        logging.warning('crear_reporte_medico error: %s', reporte.get('error'))
        flash('No se pudo crear el reporte médico en la base de datos.', 'danger')
        return redirect(url_for('upload'))

    session['last_analysis'] = json.dumps({
        'patient': patient_para_pdf,
        'study': scan_preview['study'],
        'analisis': scan_preview['analisis'],
        'summary': scan_preview['summary'],
        'training_history': scan_preview.get('training_history', []),
        'notas_clinicas': notas_clinicas,
    }, ensure_ascii=False)
    session['last_cedula'] = cedula
    session['last_correo'] = usuario.get('correo', '')
    flash('Escaneo guardado correctamente en la base de datos.', 'success')
    return redirect(url_for('resultados'))


@app.route('/upload/cancelar')
def cancelar_upload():
    scan_preview = session.pop('scan_preview', None)
    if scan_preview:
        limpiar_archivos_temporales([item['temp_path'] for item in scan_preview.get('temp_images', [])])
    flash('Escaneo cancelado. Puedes comenzar un nuevo análisis.', 'info')
    return redirect(url_for('upload'))


@app.route('/obtener_registro/<cedula>', methods=['GET'])
def obtener_registro(cedula):
    usuario = get_usuario_por_cedula(cedula)
    if not usuario or (isinstance(usuario, dict) and usuario.get('error')):
        return jsonify({'error': 'La cédula ingresada no coincide con ninguna en nuestro registro'}), 404
    paciente = get_paciente_por_cedula(cedula) or {}
    return jsonify({'usuario': usuario, 'paciente': paciente})


@app.route('/registro/editar', methods=['POST'])
def editar_registro():
    cedula = request.form.get('cedula', '').strip()
    correo = request.form.get('correo', '').strip()
    nombre_completo = request.form.get('nombre_completo', '').strip()
    edad = request.form.get('edad', '').strip()
    telefono = request.form.get('telefono', '').strip()
    direccion = request.form.get('direccion', '').strip()
    fecha_nacimiento = request.form.get('fecha_nacimiento', '').strip() or None
    sexo = request.form.get('sexo', 'F').strip() or 'F'
    contrasena = request.form.get('contrasena', '').strip()
    rol = request.form.get('rol', 'PACIENTE').strip()
    esta_activo = request.form.get('esta_activo')

    if not cedula or not correo:
        flash('Cédula y correo son obligatorios.', 'warning')
        return redirect(url_for('registro'))

    usuario_payload = {
        'nombre_completo': nombre_completo,
        'correo': correo,
        'edad': int(edad) if edad.isdigit() else None,
        'telefono': telefono,
        'direccion': direccion,
        'rol': rol,
    }
    if contrasena:
        usuario_payload['contrasena_hash'] = generate_password_hash(contrasena)
    if esta_activo is not None:
        usuario_payload['esta_activo'] = str(esta_activo).lower() in ['1', 'true', 'on', 'si', 's']

    actualizar_usuario(cedula, usuario_payload)

    paciente_payload = {
        'cedula': cedula,
        'fecha_nacimiento': fecha_nacimiento,
        'sexo': sexo if sexo in ('O', 'F', 'M') else 'F',
    }
    if get_paciente_por_cedula(cedula):
        actualizar_paciente(cedula, paciente_payload)
    else:
        crear_paciente(paciente_payload)

    session['last_correo'] = correo
    session['last_cedula'] = cedula
    flash('Registro actualizado correctamente.', 'success')
    return redirect(url_for('registro'))


@app.route('/resultados')
@login_required
def resultados():
    consulta = request.args.get('consulta', '').strip()
    registros = []
    resultado = None
    if consulta:
        raw_records = obtener_reportes_medicos(cedula=consulta, nombre=consulta, limit=100)
        for registro_item in raw_records:
            estudio = None
            prediccion = None
            if isinstance(registro_item.get('estudios'), list) and registro_item.get('estudios'):
                estudio = registro_item['estudios'][0]
            elif isinstance(registro_item.get('estudios'), dict):
                estudio = registro_item['estudios']
            if isinstance(registro_item.get('predicciones_ia'), list) and registro_item.get('predicciones_ia'):
                prediccion = registro_item['predicciones_ia'][0]
            elif isinstance(registro_item.get('predicciones_ia'), dict):
                prediccion = registro_item['predicciones_ia']

            registros.append({
                'creado_en': registro_item.get('creado_en') or registro_item.get('created_at', ''),
                'nombre_completo': estudio.get('nombre_completo') if estudio else registro_item.get('nombre_completo') or '',
                'cedula_paciente': estudio.get('cedula_paciente') if estudio else registro_item.get('cedula_paciente') or '',
                'uid_instancia_estudio': estudio.get('uid_instancia_estudio') if estudio else registro_item.get('uid_instancia_estudio') or '',
                'patologia_predicha': prediccion.get('patologia_predicha') if prediccion else registro_item.get('patologia_predicha') or '',
                'probabilidad_malignidad': prediccion.get('probabilidad_malignidad') if prediccion is not None else registro_item.get('probabilidad_malignidad'),
            })
    else:
        raw = session.get('last_analysis')
        resultado = json.loads(raw) if raw else None
    return render_template('results.html', active='results', consulta=consulta, registros=registros, resultado=resultado)


@app.route('/history/entrenar', methods=['POST'])
def history_train():
    json_file = request.files.get('json_file')
    if not json_file:
        flash('Debe seleccionar un archivo JSON.', 'warning')
        return redirect(url_for('history'))
    filename = secure_filename(json_file.filename)
    content = json_file.read()
    try:
        parsed = json.loads(content)
    except Exception:
        flash('El archivo JSON no es válido.', 'danger')
        return redirect(url_for('history'))
    saved_path = guardar_archivo_local(content, 'json', 'json')

    try:
        from agents import tasks
        if rq_queue is not None:
            rq_queue.enqueue('agents.tasks.train_job', saved_path, parsed, filename, session.get('last_cedula'))
        else:
            thread = threading.Thread(target=tasks.train_job, args=(saved_path, parsed, filename, session.get('last_cedula')), daemon=True)
            thread.start()
    except Exception:
        logging.exception('No se pudo iniciar job de entrenamiento en background')
    flash('Archivo subido. Entrenamiento en segundo plano iniciado.', 'info')
    return redirect(url_for('history'))


@app.route('/download-pdf')
@login_required
def download_pdf():
    raw = session.get('last_analysis')
    if not raw:
        flash('No hay resultados disponibles para descargar.', 'warning')
        return redirect(url_for('resultados'))

    resultado = json.loads(raw)
    pdf_bytes = generar_reporte_pdf(resultado, resultado.get('patient', {}))
    return send_file(
        io.BytesIO(pdf_bytes),
        download_name=f"reporte_mammoscan_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.pdf",
        mimetype='application/pdf',
        as_attachment=True,
    )


@app.route('/download-json')
@login_required
@medico_required
def download_json():
    raw = session.get('last_analysis')
    if not raw:
        flash('No hay resultados disponibles para descargar.', 'warning')
        return redirect(url_for('resultados'))

    resultado = json.loads(raw)
    cedula = resultado.get('patient', {}).get('cedula', 'paciente')
    json_path = guardar_analisis_json(resultado, cedula)
    if not os.path.exists(json_path):
        flash('No se pudo generar el archivo JSON.', 'danger')
        return redirect(url_for('resultados'))

    return send_file(
        json_path,
        download_name=f"entrenamiento_mammoscan_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.json",
        mimetype='application/json',
        as_attachment=True,
    )


@app.route('/download-excel')
@login_required
@medico_required
def download_excel():
    raw = session.get('last_analysis')
    if not raw:
        flash('No hay resultados disponibles para descargar.', 'warning')
        return redirect(url_for('resultados'))

    resultado = json.loads(raw)
    patient   = resultado.get('patient', {})
    study     = resultado.get('study', {})
    analisis  = resultado.get('analisis', [])
    summary   = resultado.get('summary', {})

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    AZUL  = '0F52BA'
    AZUL2 = 'E8F0FE'
    GRIS  = 'F8FAFC'
    BLANCO = 'FFFFFF'
    ROJO   = 'FEE2E2'
    VERDE  = 'DCFCE7'

    def hdr_style(ws, row, col, value, bg=AZUL, fg='FFFFFF', bold=True, size=10):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(name='Calibri', bold=bold, color=fg, size=size)
        c.fill = PatternFill('solid', fgColor=bg)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin = Side(style='thin', color='CBD5E1')
        c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        return c

    def cell_style(ws, row, col, value, bg=BLANCO, bold=False):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(name='Calibri', bold=bold, size=9)
        c.fill = PatternFill('solid', fgColor=bg)
        c.alignment = Alignment(vertical='center', wrap_text=True)
        thin = Side(style='thin', color='E2E8F0')
        c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        return c

    def auto_width(ws):
        for col in ws.columns:
            max_len = max((len(str(c.value or '')) for c in col), default=0)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(max_len + 3, 12), 50)

    wb = Workbook()

    # ----- Hoja 1: Resumen Paciente -----
    ws1 = wb.active
    ws1.title = 'Resumen Paciente'
    ws1.row_dimensions[1].height = 22

    for ci, h in enumerate(['Campo', 'Valor'], 1):
        hdr_style(ws1, 1, ci, h)
    filas_res = [
        ('Cedula',                  patient.get('cedula','')),
        ('Nombre completo',         patient.get('nombre_completo','')),
        ('Correo',                  patient.get('correo','')),
        ('Sexo',                    patient.get('sexo','')),
        ('Fecha de nacimiento',     patient.get('fecha_nacimiento','')),
        ('Descripcion estudio',     study.get('descripcion','')),
        ('UID estudio',             study.get('uid_instancia_estudio','')),
        ('Fecha estudio',           study.get('fecha_estudio','')),
        ('Total imagenes',          summary.get('total_estudios',0)),
        ('Riesgo promedio (%)',      summary.get('riesgo_promedio',0)),
        ('Probabilidad maxima (%)', summary.get('probabilidad_maxima',0)),
        ('Probabilidad minima (%)', summary.get('probabilidad_minima',0)),
        ('Detecciones alta prob.',  summary.get('detecciones_alta_probabilidad',0)),
        ('Notas clinicas',          resultado.get('notas_clinicas','')),
        ('Fecha exportacion',       datetime.utcnow().isoformat()),
    ]
    for ri, (k, v) in enumerate(filas_res, start=2):
        bg = AZUL2 if ri % 2 == 0 else GRIS
        cell_style(ws1, ri, 1, k, bg=AZUL2, bold=True)
        cell_style(ws1, ri, 2, v, bg=bg)
    auto_width(ws1)

    # ----- Hoja 2: ResNet-50 -----
    ws2 = wb.create_sheet('ResNet-50')
    hdrs2 = ['# Imagen', 'Tipo', 'Probabilidad (%)', 'Contribucion ensemble', 'Tiempo (s)', 'Score riesgo (%)']
    for ci, h in enumerate(hdrs2, 1):
        hdr_style(ws2, 1, ci, h)
    for ri, item in enumerate(analisis, start=2):
        ag = next((a for a in item.get('detalles_agentes',[]) if a.get('agente')=='ResNet'), {})
        prob = round(float(ag.get('probabilidad',0))*100, 2) if ag else 0
        bg = ROJO if item.get('score_riesgo',0) >= 50 else VERDE if item.get('score_riesgo',0) < 20 else GRIS
        for ci, v in enumerate([ri-1, item.get('tipo',''), prob,
                                 round(prob*0.40/100,4), ag.get('tiempo_seg',0), item.get('score_riesgo',0)], 1):
            cell_style(ws2, ri, ci, v, bg=bg)
    auto_width(ws2)

    # ----- Hoja 3: CNN -----
    ws3 = wb.create_sheet('CNN')
    hdrs3 = ['# Imagen', 'Tipo', 'Probabilidad (%)', 'Contribucion ensemble', 'Num. Cajas', 'Tiempo (s)', 'Score riesgo (%)']
    for ci, h in enumerate(hdrs3, 1):
        hdr_style(ws3, 1, ci, h)
    for ri, item in enumerate(analisis, start=2):
        ag = next((a for a in item.get('detalles_agentes',[]) if a.get('agente')=='CNN'), {})
        prob  = round(float(ag.get('probabilidad',0))*100, 2) if ag else 0
        boxes = len(ag.get('boxes',[])) if ag else 0
        bg = ROJO if item.get('score_riesgo',0) >= 50 else VERDE if item.get('score_riesgo',0) < 20 else GRIS
        for ci, v in enumerate([ri-1, item.get('tipo',''), prob,
                                 round(prob*0.35/100,4), boxes, ag.get('tiempo_seg',0), item.get('score_riesgo',0)], 1):
            cell_style(ws3, ri, ci, v, bg=bg)
    auto_width(ws3)

    # ----- Hoja 4: KMeans -----
    ws4 = wb.create_sheet('KMeans')
    hdrs4 = ['# Imagen', 'Tipo', 'Clusters detectados', 'Entropia', 'Inercia', 'Contribucion ensemble', 'Tiempo (s)', 'Score riesgo (%)']
    for ci, h in enumerate(hdrs4, 1):
        hdr_style(ws4, 1, ci, h)
    for ri, item in enumerate(analisis, start=2):
        ag = next((a for a in item.get('detalles_agentes',[]) if a.get('agente')=='KMeans'), {})
        clusters = ag.get('clusters_detectados',0) if ag else 0
        ent      = round(float(ag.get('entropia',0)),4) if ag else 0
        iner     = round(float(ag.get('inercia',0)),2) if ag else 0
        bg = ROJO if item.get('score_riesgo',0) >= 50 else VERDE if item.get('score_riesgo',0) < 20 else GRIS
        for ci, v in enumerate([ri-1, item.get('tipo',''), clusters, ent, iner,
                                 round(float(clusters)/max(clusters,1)*0.10,4),
                                 ag.get('tiempo_seg',0), item.get('score_riesgo',0)], 1):
            cell_style(ws4, ri, ci, v, bg=bg)
    auto_width(ws4)

    # ----- Hoja 5: IsolationForest -----
    ws5 = wb.create_sheet('IsolationForest')
    hdrs5 = ['# Imagen', 'Tipo', 'Anomalias detectadas', 'Score anomalia', 'Contribucion ensemble', 'Tiempo (s)', 'Score riesgo (%)']
    for ci, h in enumerate(hdrs5, 1):
        hdr_style(ws5, 1, ci, h)
    for ri, item in enumerate(analisis, start=2):
        ag = next((a for a in item.get('detalles_agentes',[]) if a.get('agente')=='IsolationForest'), {})
        n_an  = ag.get('anomalias_detectadas',0) if ag else 0
        score = round(float(ag.get('score_anomalia',0)),4) if ag else 0
        bg = ROJO if item.get('score_riesgo',0) >= 50 else VERDE if item.get('score_riesgo',0) < 20 else GRIS
        for ci, v in enumerate([ri-1, item.get('tipo',''), n_an, score,
                                 round(float(n_an)/10*0.15,4),
                                 ag.get('tiempo_seg',0), item.get('score_riesgo',0)], 1):
            cell_style(ws5, ri, ci, v, bg=bg)
    auto_width(ws5)

    # ----- Hoja 6: Ensemble -----
    ws6 = wb.create_sheet('Ensemble')
    hdrs6 = ['# Imagen', 'Tipo', 'Score ResNet (40%)', 'Score CNN (35%)', 'Score IsoForest (15%)', 'Score KMeans (10%)', 'Score final (%)', 'Nivel riesgo']
    for ci, h in enumerate(hdrs6, 1):
        hdr_style(ws6, 1, ci, h)
    for ri, item in enumerate(analisis, start=2):
        ags    = {a.get('agente'): a for a in item.get('detalles_agentes',[])}
        r_prob = float(ags.get('ResNet',{}).get('probabilidad',0))
        c_prob = float(ags.get('CNN',{}).get('probabilidad',0))
        n_an   = float(ags.get('IsolationForest',{}).get('anomalias_detectadas',0))
        clust  = float(ags.get('KMeans',{}).get('clusters_detectados',0))
        bg = ROJO if item.get('score_riesgo',0) >= 50 else VERDE if item.get('score_riesgo',0) < 20 else GRIS
        for ci, v in enumerate([
            ri-1, item.get('tipo',''),
            round(r_prob*0.40,4), round(c_prob*0.35,4),
            round((n_an/10)*0.15,4), round((clust/max(clust,1))*0.10,4),
            item.get('score_riesgo',0), item.get('nivel_riesgo','')
        ], 1):
            cell_style(ws6, ri, ci, v, bg=bg)
    auto_width(ws6)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        download_name='resultados_dermaia_' + datetime.utcnow().strftime('%Y%m%d_%H%M%S') + '.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
    )

@app.route('/download-history')
def download_history():
    records = obtener_importaciones_excel(limit=100)
    if not records:
        flash('No hay historial para descargar.', 'warning')
        return redirect(url_for('history'))

    summary_data = format_history_records(records)
    history_rows = []
    for record in records:
        meta = record.get('import_metadata') or record.get('metadatos_importacion') or {}
        training = meta.get('training_history', [])
        run_label = record.get('file_name') or record.get('nombre_archivo', '')
        for epoch in training:
            history_rows.append({
                'Run': run_label,
                'Epoch': epoch.get('epoch'),
                'loss': epoch.get('loss'),
                'accuracy': epoch.get('accuracy'),
                'val_loss': epoch.get('val_loss'),
                'val_accuracy': epoch.get('val_accuracy'),
            })

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        pd.DataFrame(summary_data).to_excel(writer, index=False, sheet_name='Resumen')
        if history_rows:
            pd.DataFrame(history_rows).to_excel(writer, index=False, sheet_name='Entrenamiento')
    output.seek(0)

    return send_file(
        output,
        download_name='history.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
    )


@app.route('/train_model', methods=['GET', 'POST'])
@login_required
@medico_required
def train_model():
    if request.method == 'POST':
        json_file = request.files.get('json_file')
        if json_file:
            filename = secure_filename(json_file.filename)
            if not filename.lower().endswith('.json'):
                flash('El archivo debe tener extensión .json.', 'warning')
                return redirect(url_for('train_model'))
            content = json_file.read()
            try:
                parsed = json.loads(content)
            except Exception:
                flash('El archivo JSON no es válido.', 'danger')
                return redirect(url_for('train_model'))
            saved_path = guardar_archivo_local(content, 'json', 'json')
            try:
                from agents import tasks
                if rq_queue is not None:
                    rq_queue.enqueue('agents.tasks.train_job', saved_path, parsed, filename, session.get('last_cedula'))
                else:
                    thread = threading.Thread(target=tasks.train_job, args=(saved_path, parsed, filename, session.get('last_cedula')), daemon=True)
                    thread.start()
            except Exception:
                logging.exception('No se pudo iniciar job de entrenamiento en background')
            flash('Archivo subido. Entrenamiento en segundo plano iniciado.', 'info')
            return redirect(url_for('train_model'))
    return render_template('train.html', active='train')




@app.route('/api/resultados')
def api_resultados():
    cedula = request.args.get('cedula', '').strip()
    if not cedula:
        return jsonify([])
    reportes = obtener_reportes_completos_por_cedula(cedula, limit=100)
    estudio = obtener_estudio_por_cedula(cedula) or {}
    lineas = obtener_lineas_por_cedula(cedula, limit=100)
    result = []
    for r in reportes:
        pred = {}
        if isinstance(r.get('predicciones_ia'), list) and r['predicciones_ia']:
            pred = r['predicciones_ia'][0]
        elif isinstance(r.get('predicciones_ia'), dict):
            pred = r['predicciones_ia']
        linea_id = pred.get('linea_analisis_id', '')
        linea = next((l for l in lineas if l.get('id') == linea_id), {})
        result.append({
            'cedula_paciente': r.get('cedula_paciente', ''),
            'fecha_estudio': estudio.get('fecha_estudio', ''),
            'reporte': {
                'id': r.get('id', ''),
                'notas_clinicas': r.get('notas_clinicas', ''),
                'ruta_pdf_almacenamiento': r.get('ruta_pdf_almacenamiento', ''),
                'esta_verificado': r.get('esta_verificado', False),
                'creado_en': r.get('creado_en', ''),
            },
            'prediccion': {
                'id': pred.get('id', ''),
                'modelo_nombre': pred.get('modelo_nombre', ''),
                'patologia_predicha': pred.get('patologia_predicha', ''),
                'probabilidad_malignidad': pred.get('probabilidad_malignidad'),
            },
            'estudio': {
                'uid_instancia_estudio': estudio.get('uid_instancia_estudio', ''),
                'fecha_estudio': estudio.get('fecha_estudio', ''),
                'descripcion': estudio.get('descripcion', ''),
            },
            'linea': {
                'id': linea.get('id', ''),
                'estado': linea.get('estado', ''),
                'tiempo_ejecucion_segundos': linea.get('tiempo_ejecucion_segundos'),
                'parametros_filtros_reduccion_ruido': linea.get('parametros_filtros_reduccion_ruido', {}),
                'registros_preprocesamiento': linea.get('registros_preprocesamiento', ''),
            },
        })
    return jsonify(result)


@app.route('/api/detalle_resultado')
def api_detalle_resultado():
    reporte_id = request.args.get('reporte_id', '').strip()
    pred_id = request.args.get('pred_id', '').strip()
    reporte = {}
    pred = {}
    linea = {}
    imagen = {}
    cedula = ''
    if reporte_id:
        reporte = obtener_reporte_medico_por_id(reporte_id) or {}
        cedula = reporte.get('cedula_paciente', '')
        pred_id_r = reporte.get('prediccion_id', pred_id)
        if pred_id_r:
            pred = obtener_prediccion_por_id(pred_id_r) or {}
    elif pred_id:
        pred = obtener_prediccion_por_id(pred_id) or {}
        cedula = pred.get('cedula_paciente', '')
    linea_id = pred.get('linea_analisis_id', '')
    if linea_id:
        linea = obtener_linea_por_id(linea_id) or {}
        imagen_id = linea.get('imagen_id', '')
        if imagen_id:
            imagen = obtener_imagen_por_id(imagen_id) or {}
    estudio = obtener_estudio_por_cedula(cedula) or {} if cedula else {}
    paciente = get_paciente_por_cedula(cedula) or {} if cedula else {}
    usuario = get_usuario_por_cedula(cedula) or {} if cedula else {}
    return jsonify({
        'reporte': reporte,
        'prediccion': pred,
        'linea': linea,
        'imagen': imagen,
        'estudio': estudio,
        'paciente': paciente,
        'usuario': usuario,
    })


@app.route('/api/descargar_pdf')
@login_required
def api_descargar_pdf():
    reporte_id = request.args.get('reporte_id', '').strip()
    if not reporte_id:
        return jsonify({'error': 'reporte_id requerido'}), 400
    reporte = obtener_reporte_medico_por_id(reporte_id) or {}
    if not reporte:
        return jsonify({'error': 'Reporte no encontrado'}), 404
    cedula_reporte = reporte.get('cedula_paciente', '')
    usuario_sesion = session.get('usuario', {})
    if usuario_sesion.get('rol') == 'PACIENTE' and cedula_reporte != usuario_sesion.get('cedula'):
        return jsonify({'error': 'Sin autorización'}), 403
    ruta = reporte.get('ruta_pdf_almacenamiento', '')
    if ruta and os.path.exists(ruta):
        return send_file(
            ruta,
            download_name=os.path.basename(ruta),
            mimetype='application/pdf',
            as_attachment=True,
        )
    pred_id = reporte.get('prediccion_id', '')
    pred = obtener_prediccion_por_id(pred_id) or {} if pred_id else {}
    linea_id = pred.get('linea_analisis_id', '')
    linea = obtener_linea_por_id(linea_id) or {} if linea_id else {}
    imagen_id = linea.get('imagen_id', '')
    imagen = obtener_imagen_por_id(imagen_id) or {} if imagen_id else {}
    estudio = obtener_estudio_por_cedula(cedula_reporte) or {} if cedula_reporte else {}
    paciente = get_paciente_por_cedula(cedula_reporte) or {} if cedula_reporte else {}
    usuario_data = get_usuario_por_cedula(cedula_reporte) or {} if cedula_reporte else {}
    resultado_pdf = {
        'patient': {
            'cedula': cedula_reporte,
            'nombre_completo': usuario_data.get('nombre_completo', ''),
            'correo': usuario_data.get('correo', ''),
            'sexo': paciente.get('sexo', ''),
            'fecha_nacimiento': paciente.get('fecha_nacimiento', ''),
        },
        'imagenes': [{
            'nombre': imagen.get('nombre_archivo', 'imagen'),
            'patologia': pred.get('patologia_predicha', 'N/D'),
            'probabilidad_malignidad': pred.get('probabilidad_malignidad', 0),
            'agentes': {
                'ResNet': pred.get('prob_resnet', 0),
                'CNN': pred.get('prob_cnn', 0),
                'IsolationForest': pred.get('score_isolation_forest', 0),
                'KMeans': pred.get('score_kmeans', 0),
            }
        }],
        'notas_clinicas': reporte.get('notas_clinicas', ''),
    }
    pdf_bytes = generar_reporte_pdf(resultado_pdf, resultado_pdf['patient'])
    nombre_pdf = f"reporte_{cedula_reporte}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.pdf"
    return send_file(
        io.BytesIO(pdf_bytes),
        download_name=nombre_pdf,
        mimetype='application/pdf',
        as_attachment=True,
    )

@app.route('/api/eliminar_resultado', methods=['POST'])
def api_eliminar_resultado():
    data = request.get_json() or {}
    reporte_id = (data.get('reporte_id') or '').strip()
    if not reporte_id:
        return jsonify({'ok': False, 'error': 'reporte_id requerido'})
    resultado = eliminar_resultado_completo(reporte_id)
    if isinstance(resultado, dict) and resultado.get('error'):
        return jsonify({'ok': False, 'error': resultado['error']})
    return jsonify({'ok': True})


@app.route('/api/archivos_locales')
def api_archivos_locales():
    import datetime as dt_mod
    reports_dir = os.path.join(os.getcwd(), 'uploads', 'reports')
    archivos = []
    try:
        if os.path.exists(reports_dir):
            for nombre in sorted(os.listdir(reports_dir)):
                ruta = os.path.join(reports_dir, nombre)
                if os.path.isfile(ruta):
                    stat = os.stat(ruta)
                    kb = round(stat.st_size / 1024, 1)
                    fecha = dt_mod.datetime.fromtimestamp(stat.st_mtime).isoformat()
                    archivos.append({'nombre': nombre, 'tamanio': str(kb) + ' KB', 'fecha': fecha})
    except Exception:
        pass
    archivos.sort(key=lambda a: a.get('fecha', ''), reverse=True)
    return jsonify(archivos)


@app.route('/api/eliminar_archivo_local', methods=['POST'])
def api_eliminar_archivo_local():
    data = request.get_json() or {}
    nombre = (data.get('nombre') or '').strip()
    if not nombre or '/' in nombre or '\\' in nombre or '..' in nombre:
        return jsonify({'ok': False, 'error': 'Nombre de archivo invalido'})
    ruta = os.path.join(os.getcwd(), 'uploads', 'reports', nombre)
    if not os.path.exists(ruta):
        return jsonify({'ok': False, 'error': 'Archivo no encontrado'})
    try:
        os.remove(ruta)
        return jsonify({'ok': True})
    except Exception as exc:
        return jsonify({'ok': False, 'error': str(exc)})


@app.route('/api/historial_entrenamiento')
def api_historial_entrenamiento():
    importaciones = obtener_importaciones_excel(limit=50)
    agentes_nombres = ['ResNet', 'CNN', 'KMeans', 'IsolationForest', 'Keras']
    por_agente = {a: [] for a in agentes_nombres}
    for imp in importaciones:
        meta = imp.get('metadatos_importacion') or imp.get('import_metadata') or {}
        if isinstance(meta, str):
            try:
                meta = json.loads(meta)
            except Exception:
                meta = {}
        fname = imp.get('nombre_archivo') or imp.get('file_name', '')
        fecha_imp = imp.get('importado_en') or imp.get('imported_at', '')
        for epoch in meta.get('training_history', []):
            por_agente['Keras'].append({
                'run': fname,
                'epoch': epoch.get('epoch'),
                'perdida': epoch.get('loss'),
                'accuracy': epoch.get('accuracy'),
                'val_loss': epoch.get('val_loss'),
                'val_accuracy': epoch.get('val_accuracy'),
                'fecha': fecha_imp,
            })
        for muestra in meta.get('agentes_escaneo_historial', []):
            vals = muestra.get('valores_agentes', {})
            for agente in ['ResNet', 'CNN', 'KMeans', 'IsolationForest']:
                v = vals.get(agente)
                if v is not None and v != 'No se aplicó':
                    por_agente[agente].append({
                        'run': fname,
                        'id_muestra': muestra.get('id_muestra'),
                        'label': muestra.get('label', ''),
                        'valor': v,
                        'probabilidad_malignidad': muestra.get('probabilidad_malignidad'),
                        'fecha': fecha_imp,
                    })
    return jsonify({'historial': por_agente, 'agentes': agentes_nombres})

@app.route('/api/json_entrenamientos')
def api_json_entrenamientos():
    registros = obtener_importaciones_excel(limit=200)
    result = []
    for r in registros:
        result.append({
            'id': r.get('id', ''),
            'nombre': r.get('nombre_archivo', ''),
            'total_registros': r.get('total_registros_entrenamiento', 0),
            'fecha': r.get('importado_en', ''),
        })
    return jsonify(result)


@app.route('/api/eliminar_json_entrenamiento', methods=['POST'])
def api_eliminar_json_entrenamiento():
    data = request.get_json() or {}
    importacion_id = (data.get('id') or '').strip()
    if not importacion_id:
        return jsonify({'ok': False, 'error': 'id requerido'})
    res = eliminar_importacion_excel(importacion_id)
    if isinstance(res, dict) and res.get('error'):
        return jsonify({'ok': False, 'error': res['error']})
    return jsonify({'ok': True})


@app.route('/imagenes')
@login_required
@medico_required
def imagenes():
    return render_template('imagenes.html', active='imagenes')


@app.route('/api/imagenes_locales')
@login_required
def api_imagenes_locales():
    import datetime as dt_mod
    images_dir = os.path.join(os.getcwd(), 'uploads', 'images')
    archivos = []
    try:
        if os.path.exists(images_dir):
            for nombre in sorted(os.listdir(images_dir)):
                ruta = os.path.join(images_dir, nombre)
                if os.path.isfile(ruta):
                    stat = os.stat(ruta)
                    kb   = round(stat.st_size / 1024, 1)
                    fecha = dt_mod.datetime.fromtimestamp(stat.st_mtime).isoformat()
                    archivos.append({'nombre': nombre, 'tamanio': str(kb) + ' KB', 'fecha': fecha})
    except Exception:
        pass
    archivos.sort(key=lambda a: a.get('fecha', ''), reverse=True)
    return jsonify(archivos)


@app.route('/api/eliminar_imagen_local', methods=['POST'])
@login_required
@medico_required
def api_eliminar_imagen_local():
    data = request.get_json() or {}
    nombre = (data.get('nombre') or '').strip()
    if not nombre or '/' in nombre or '\\' in nombre or '..' in nombre:
        return jsonify({'ok': False, 'error': 'Nombre invalido'})
    ruta = os.path.join(os.getcwd(), 'uploads', 'images', nombre)
    if not os.path.exists(ruta):
        return jsonify({'ok': False, 'error': 'Archivo no encontrado'})
    try:
        os.remove(ruta)
        return jsonify({'ok': True})
    except Exception as exc:
        return jsonify({'ok': False, 'error': str(exc)})


@app.route('/api/enviar_correo', methods=['POST'])
@login_required
def api_enviar_correo():
    data = request.get_json() or {}
    reporte_id = data.get('reporte_id', '').strip()
    pred_id = data.get('pred_id', '').strip()
    correo_destino = data.get('correo_destino', '').strip()

    reporte = obtener_reporte_medico_por_id(reporte_id) if reporte_id else {}
    cedula = reporte.get('cedula_paciente', '') if reporte else ''
    pred = obtener_prediccion_por_id(pred_id) if pred_id else {}
    linea_id = pred.get('linea_analisis_id', '') if pred else ''
    linea = obtener_linea_por_id(linea_id) if linea_id else {}
    imagen_id = linea.get('imagen_id', '') if linea else ''
    imagen = obtener_imagen_por_id(imagen_id) if imagen_id else {}
    estudio = obtener_estudio_por_cedula(cedula) if cedula else {}
    paciente = get_paciente_por_cedula(cedula) if cedula else {}
    usuario = get_usuario_por_cedula(cedula) if cedula else {}

    if not correo_destino and usuario and isinstance(usuario, dict):
        correo_destino = usuario.get('correo', '')
    if not correo_destino:
        return jsonify({'ok': False, 'error': 'No se encontró correo destino.'})

    resultado_mock = {
        'patient': {
            'cedula': cedula,
            'nombre_completo': usuario.get('nombre_completo', '') if usuario else '',
            'correo': correo_destino,
            'sexo': paciente.get('sexo', '') if paciente else '',
            'fecha_nacimiento': paciente.get('fecha_nacimiento', '') if paciente else '',
        },
        'study': estudio or {},
        'notas_clinicas': reporte.get('notas_clinicas', '') if reporte else '',
        'summary': {},
        'analisis': [],
    }
    pdf_bytes = generar_reporte_pdf(resultado_mock, resultado_mock['patient'])

    SENDER = 'unisimoestudiante@gmail.com'
    APP_PASS = os.getenv('GMAIL_APP_PASSWORD', '')
    if not APP_PASS:
        return jsonify({'ok': False, 'error': 'Contraseña SMTP no configurada (GMAIL_APP_PASSWORD).'})

    msg = MIMEMultipart()
    msg['From'] = SENDER
    msg['To'] = correo_destino
    msg['Subject'] = 'Reporte DermaIAInsight – Resultado de análisis'
    nombre_paciente = resultado_mock['patient'].get('nombre_completo', 'Paciente')
    lineas_cuerpo = [
        'Estimado/a ' + nombre_paciente + ',',
        '',
        'Adjunto encontrara el reporte de su analisis medico generado por la plataforma DermaIAInsight.',
        '',
        'Este informe contiene los resultados del analisis de imagenes realizado con los agentes de IA.',
        'Los resultados son orientativos y deben ser confirmados por un especialista medico.',
        '',
        'DermaIAInsight - Universidad Simon Bolivar, sede Cucuta',
        'Ingenieria de Sistemas',
    ]
    cuerpo = '\n'.join(lineas_cuerpo)
    msg.attach(MIMEText(cuerpo, 'plain', 'utf-8'))
    part = MIMEBase('application', 'octet-stream')
    part.set_payload(pdf_bytes)
    encoders.encode_base64(part)
    fname = f"reporte_dermaia_{cedula}_{datetime.utcnow().strftime('%Y%m%d')}.pdf"
    part.add_header('Content-Disposition', f'attachment; filename="{fname}"')
    msg.attach(part)

    try:
        with smtplib.SMTP('smtp.gmail.com', 587, timeout=15) as server:
            server.ehlo()
            server.starttls()
            server.login(SENDER, APP_PASS)
            server.sendmail(SENDER, correo_destino, msg.as_bytes())
        return jsonify({'ok': True, 'correo': correo_destino})
    except Exception as exc:
        logging.exception('Error al enviar correo')
        return jsonify({'ok': False, 'error': str(exc)})



@app.route('/api/listar_usuarios')
@login_required
@medico_required
def api_listar_usuarios():
    cedula_filtro = request.args.get('cedula', '').strip()
    usuarios = listar_todos_usuarios(limit=300)
    if cedula_filtro:
        usuarios = [u for u in usuarios if cedula_filtro.lower() in (u.get('cedula') or '').lower()]
    return jsonify(usuarios)


@app.route('/api/cambiar_rol_usuario', methods=['POST'])
@login_required
@medico_required
def api_cambiar_rol_usuario():
    data = request.get_json() or {}
    cedula    = (data.get('cedula') or '').strip()
    nuevo_rol = (data.get('rol') or '').strip().upper()
    if not cedula or nuevo_rol not in ('PACIENTE', 'MEDICO'):
        return jsonify({'ok': False, 'error': 'cedula y rol (PACIENTE|MEDICO) son requeridos'})
    cedula_sesion = session.get('usuario', {}).get('cedula', '')
    if cedula == cedula_sesion:
        return jsonify({'ok': False, 'error': 'No puede cambiar su propio rol'})
    resultado = actualizar_rol_usuario(cedula, nuevo_rol)
    if isinstance(resultado, dict) and resultado.get('error'):
        return jsonify({'ok': False, 'error': resultado['error']})
    return jsonify({'ok': True, 'cedula': cedula, 'rol': nuevo_rol})


@app.route('/api/eliminar_usuario', methods=['POST'])
@login_required
@medico_required
def api_eliminar_usuario():
    data = request.get_json() or {}
    cedula = (data.get('cedula') or '').strip()
    if not cedula:
        return jsonify({'ok': False, 'error': 'cedula requerida'})
    cedula_sesion = session.get('usuario', {}).get('cedula', '')
    if cedula == cedula_sesion:
        return jsonify({'ok': False, 'error': 'No puede eliminar su propia cuenta'})
    resultado = eliminar_usuario_completo(cedula)
    if isinstance(resultado, dict) and resultado.get('error'):
        return jsonify({'ok': False, 'error': resultado['error']})
    return jsonify({'ok': True})


if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=7777)
